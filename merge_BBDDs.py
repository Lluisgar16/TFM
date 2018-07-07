# -*- coding: cp1252 -*-.
import sys
import os,glob
import unicodedata
import openpyxl as xls

def LevenshteinDistance(s1, s2):
    if len(s1) < len(s2):
        return LevenshteinDistance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1 
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    return previous_row[-1]

def Similarity(a, b):
    distance = LevenshteinDistance(a,b)
    return (max(len(a),len(b))-distance)/float(max(len(a),len(b)))

def NormalizeString(string):
    try:
        return unicodedata.normalize('NFC',string)
    except TypeError:
        return ''

def main():
    print sys.version
    os.chdir(r"/Users/lluisgar16/Desktop/TFM/DATA")

    workbook = xls.load_workbook("BBDD_Preguntas_orales.xlsx")
    diputados = xls.load_workbook("diputados 2011.xlsx")
    sheet_diputados = diputados.active
    sheet = workbook.active
    diputados_names = []
    for row in sheet_diputados.iter_rows():
        # Name, Party, Univ
        build_name = NormalizeString(row[3].value)+" "+NormalizeString(row[4].value)+", "+NormalizeString(row[2].value)
        if build_name.count("Apellido 1") != 0:
            continue
        else:
            diputados_names.append((build_name,NormalizeString(row[1].value),str(int(NormalizeString(row[9].value)!="No"))))

    csvlines = []
    csvlines.append('ID;AUTHOR;GENDER;PARTY;UNIV;CODE;SUBCODE;TITLE')
    for i,row in enumerate(sheet.iter_rows()):
        print str(i)
        name = NormalizeString(row[7].value)
        if name.count("AUTHOR") != 0:
            continue
        for auxname in diputados_names:
            if Similarity(name,auxname[0]) >= 0.8:
                print auxname[0],name, Similarity(name,auxname[0])
                # raw_input()
                line = str(row[0].value)+';'+name+';'+\
                       str(row[8].value)+';'+auxname[1]+';'+auxname[2]+';'+\
                       str(row[15].value)+';'+str(row[16].value)+';'+\
                       NormalizeString(row[6].value)
                csvlines.append(line.encode('utf-8'))
                continue

    with open("BBDD_Preguntas_orales.csv", "wb") as csv:
        for line in csvlines:
            csv.write(line+"\n")

if __name__ == '__main__':
    main()